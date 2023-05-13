import openpyxl
import pytest


class SubmitData:
    data_form_submission = [{
                                "firstname": "Patryk",
                                "email": "test@g.com",
                                "password": "password1@",
                                "status": "option1",
                                "birth": "01081999"
                             },
                            {
                                "firstname": "Kacper",
                                "email": "test2@g.com",
                                "password": "password2@",
                                "status": "option2",
                                "birth": "02092000"
                            }]

    @staticmethod
    def exel_test_data(test_case_name):
        data = {}
        exel = openpyxl.load_workbook("/TestData\Testcases.xlsx")
        sheet = exel.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [data]
